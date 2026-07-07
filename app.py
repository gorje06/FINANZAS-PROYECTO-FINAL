"""FinanCuota — simulador de crédito vehicular. Grupo 2 — UPC SI642."""

import io
import json
import os
import re
from datetime import date
from functools import wraps

from flask import Flask, Response, flash, redirect, render_template, request, session, url_for
from werkzeug.security import check_password_hash, generate_password_hash

from db import DB_PATH, catalogo_count, get_conn, init_db
from finance import (
    MODALIDAD_COMPRA_INTELIGENTE,
    MODALIDAD_CONVENCIONAL,
    build_schedule,
    calcular_tcea_desde_tir_mensual,
    calcular_tir,
    calcular_van,
)


app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "financuota-dev-secret")

ROL_ADMIN = "admin"
ROL_USUARIO = "usuario"
ADMIN_LOGIN = "admin"
ADMIN_PASSWORD = "adminupc"
ADMIN_DNI = "00000001"

init_db()


def _ensure_admin_account() -> None:
    conn = get_conn()
    row = conn.execute(
        "SELECT id_usuario FROM usuario WHERE usuario_login = ?", (ADMIN_LOGIN,)
    ).fetchone()
    if not row:
        conn.execute(
            """
            INSERT INTO usuario (usuario_login, password_hash, dni_usuario, rol)
            VALUES (?, ?, ?, ?)
            """,
            (ADMIN_LOGIN, generate_password_hash(ADMIN_PASSWORD), ADMIN_DNI, ROL_ADMIN),
        )
        conn.commit()
    else:
        conn.execute(
            "UPDATE usuario SET rol = ? WHERE usuario_login = ?",
            (ROL_ADMIN, ADMIN_LOGIN),
        )
        conn.commit()
    conn.close()


_ensure_admin_account()


def _is_admin() -> bool:
    return session.get("rol") == ROL_ADMIN


@app.context_processor
def inject_globals():
    return {"is_admin": _is_admin()}


def login_required(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        return func(*args, **kwargs)

    return wrapper


def admin_required(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        if not _is_admin():
            flash("No tienes permisos de administrador.")
            return redirect(url_for("dashboard"))
        return func(*args, **kwargs)

    return wrapper


PERIODO_TASA_ETIQUETAS = {
    0: "Base 360 días",
    1: "Cada dos meses",
    2: "Mensual (el porcentaje ya es la tasa mensual)",
    3: "Semestral",
    4: "Cuatrimestral",
    5: "Trimestral",
    6: "Bimestral",
    7: "Anual (se convierte a tasa mensual)",
}

PERIODO_OPCIONES = [(k, v) for k, v in PERIODO_TASA_ETIQUETAS.items()]

CASOS_PRUEBA = {
    "carlos": {
        "nombres_cliente": "Carlos",
        "apellidos_cliente": "Ramírez",
        "dni_cliente": "72859461",
        "correo_cliente": "carlos.ramirez@email.com",
        "telefono_cliente": "987654321",
        "direccion_cliente": "Av. Javier Prado 123, Lima",
        "ingresos_mensuales": "8500",
        "marca_vehiculo": "Toyota",
        "modelo_vehiculo": "Corolla 2026",
        "precio_vehiculo": "85000",
        "cuota_inicial_pct": "20",
        "moneda": "Soles",
        "modalidad": MODALIDAD_CONVENCIONAL,
        "cuota_balon_pct": "0",
        "tipo_tasa": "Efectiva",
        "periodo_tasa": "7",
        "tasa_interes": "12.5",
        "capitalizacion": "12",
        "plazo_meses": "48",
        "fecha_desembolso": "2026-08-15",
        "periodo_gracia": "Parcial",
        "meses_gracia": "2",
        "seguro_desgravamen": "0.03",
        "seguro_vehicular": "0.025",
        "portes": "4",
        "gastos_notariales": "450",
        "gastos_registrales": "320",
        "costos_iniciales": "150",
        "tipo_cambio": "",
    },
    "maria": {
        "nombres_cliente": "María",
        "apellidos_cliente": "López",
        "dni_cliente": "40987654",
        "correo_cliente": "maria.lopez@email.com",
        "telefono_cliente": "912345678",
        "direccion_cliente": "Calle Los Pinos 456, Arequipa",
        "ingresos_mensuales": "12000",
        "marca_vehiculo": "Hyundai",
        "modelo_vehiculo": "Tucson 2026",
        "precio_vehiculo": "32000",
        "cuota_inicial_pct": "15",
        "moneda": "Dólares",
        "modalidad": MODALIDAD_COMPRA_INTELIGENTE,
        "cuota_balon_pct": "40",
        "tipo_tasa": "Efectiva",
        "periodo_tasa": "7",
        "tasa_interes": "11.8",
        "capitalizacion": "12",
        "plazo_meses": "36",
        "fecha_desembolso": "2026-09-01",
        "periodo_gracia": "Ninguno",
        "meses_gracia": "0",
        "seguro_desgravamen": "0.035",
        "seguro_vehicular": "0.03",
        "portes": "5",
        "gastos_notariales": "180",
        "gastos_registrales": "120",
        "costos_iniciales": "80",
        "tipo_cambio": "3.75",
    },
    "roberto": {
        "nombres_cliente": "Roberto",
        "apellidos_cliente": "Mendoza",
        "dni_cliente": "45678912",
        "correo_cliente": "roberto.mendoza@email.com",
        "telefono_cliente": "934567890",
        "direccion_cliente": "Av. La Molina 890, Lima",
        "ingresos_mensuales": "15000",
        "marca_vehiculo": "Mazda",
        "modelo_vehiculo": "CX-5 2025",
        "precio_vehiculo": "155800",
        "cuota_inicial_pct": "20",
        "moneda": "Soles",
        "modalidad": MODALIDAD_COMPRA_INTELIGENTE,
        "cuota_balon_pct": "30",
        "tipo_tasa": "Efectiva",
        "periodo_tasa": "7",
        "tasa_interes": "12.5",
        "capitalizacion": "12",
        "plazo_meses": "48",
        "fecha_desembolso": "2026-08-20",
        "periodo_gracia": "Parcial",
        "meses_gracia": "2",
        "seguro_desgravamen": "0.03",
        "seguro_vehicular": "0.025",
        "portes": "4",
        "gastos_notariales": "520",
        "gastos_registrales": "380",
        "costos_iniciales": "200",
        "tipo_cambio": "",
    },
    "lucia": {
        "nombres_cliente": "Lucía",
        "apellidos_cliente": "Fernández",
        "dni_cliente": "51890234",
        "correo_cliente": "lucia.fernandez@email.com",
        "telefono_cliente": "976543210",
        "direccion_cliente": "Calle Las Begonias 220, Miraflores",
        "ingresos_mensuales": "22000",
        "marca_vehiculo": "Ford",
        "modelo_vehiculo": "Mustang 2024",
        "precio_vehiculo": "310000",
        "cuota_inicial_pct": "15",
        "moneda": "Soles",
        "modalidad": MODALIDAD_COMPRA_INTELIGENTE,
        "cuota_balon_pct": "50",
        "tipo_tasa": "Efectiva",
        "periodo_tasa": "7",
        "tasa_interes": "11.8",
        "capitalizacion": "12",
        "plazo_meses": "36",
        "fecha_desembolso": "2026-09-10",
        "periodo_gracia": "Ninguno",
        "meses_gracia": "0",
        "seguro_desgravamen": "0.035",
        "seguro_vehicular": "0.028",
        "portes": "5",
        "gastos_notariales": "680",
        "gastos_registrales": "450",
        "costos_iniciales": "250",
        "tipo_cambio": "",
    },
}

CASOS_PRUEBA_ORDEN = [
    ("carlos", "Caso 1 — Carlos Ramírez (convencional, gracia parcial)"),
    ("maria", "Caso 2 — María López (Compra Inteligente, USD, balón 40%)"),
    ("roberto", "Caso 3 — Roberto Mendoza (Compra Inteligente, soles, balón 30%)"),
    ("lucia", "Caso 4 — Lucía Fernández (Compra Inteligente, soles, balón 50%)"),
]

CREDIT_DEFAULTS_KEYS = (
    "moneda",
    "tipo_tasa",
    "cuota_inicial_pct",
    "plazo_meses",
    "periodo_tasa",
    "capitalizacion",
    "periodo_gracia",
    "meses_gracia",
    "seguro_desgravamen",
    "seguro_vehicular",
    "portes",
    "modalidad",
    "cuota_balon_pct",
    "gastos_notariales",
    "gastos_registrales",
    "costos_iniciales",
    "tipo_cambio",
)

DEFAULTS_KEYS = CREDIT_DEFAULTS_KEYS + (
    "nombres_cliente",
    "apellidos_cliente",
    "dni_cliente",
    "correo_cliente",
    "telefono_cliente",
    "direccion_cliente",
    "ingresos_mensuales",
    "marca_vehiculo",
    "modelo_vehiculo",
    "precio_vehiculo",
    "tasa_interes",
    "fecha_desembolso",
)


def etiqueta_periodo_tasa(valor) -> str:
    try:
        clave = int(valor)
    except (TypeError, ValueError):
        clave = 7
    return PERIODO_TASA_ETIQUETAS.get(clave, PERIODO_TASA_ETIQUETAS[7])


def etiqueta_capitalizacion(valor) -> str:
    if valor is None or valor == "":
        return "—"
    try:
        c = int(valor)
    except (TypeError, ValueError):
        return str(valor)
    nombres = {1: "Una vez al año", 2: "Semestral", 4: "Trimestral", 12: "Mensual"}
    return nombres.get(c, f"{c} veces al año")


def _num(value, default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return default


def _row_get(row, key: str, default=None):
    try:
        return row[key]
    except (KeyError, TypeError, IndexError):
        return default


def _base_credit_defaults() -> dict:
    return {
        "moneda": "Soles",
        "tipo_tasa": "Efectiva",
        "cuota_inicial_pct": 20,
        "plazo_meses": 48,
        "periodo_tasa": 7,
        "capitalizacion": 12,
        "periodo_gracia": "Ninguno",
        "meses_gracia": 0,
        "seguro_desgravamen": 0.03,
        "seguro_vehicular": 0.025,
        "portes": 4,
        "modalidad": MODALIDAD_CONVENCIONAL,
        "cuota_balon_pct": 0,
        "gastos_notariales": 0,
        "gastos_registrales": 0,
        "costos_iniciales": 0,
        "tipo_cambio": 3.75,
    }


def _get_credit_defaults() -> dict:
    stored = session.get("defaults") or {}
    base = _base_credit_defaults()
    for key in CREDIT_DEFAULTS_KEYS:
        if key in stored and str(stored[key]).strip() != "":
            base[key] = stored[key]
    return base


def _get_wizard_defaults() -> dict:
    defaults = _get_credit_defaults()
    defaults.update(
        {
            "nombres_cliente": "",
            "apellidos_cliente": "",
            "dni_cliente": "",
            "correo_cliente": "",
            "telefono_cliente": "",
            "direccion_cliente": "",
            "ingresos_mensuales": "",
            "marca_vehiculo": "",
            "modelo_vehiculo": "",
            "precio_vehiculo": "",
            "tasa_interes": "",
            "fecha_desembolso": date.today().isoformat(),
        }
    )
    return defaults


def _change_password(user_id: int, current: str, new_pass: str, confirm: str) -> None:
    if len(new_pass) < 6:
        raise ValueError("La nueva contraseña debe tener al menos 6 caracteres.")
    if new_pass != confirm:
        raise ValueError("Las contraseñas nuevas no coinciden.")
    conn = get_conn()
    user = conn.execute("SELECT password_hash FROM usuario WHERE id_usuario = ?", (user_id,)).fetchone()
    if not user or not check_password_hash(user["password_hash"], current):
        conn.close()
        raise ValueError("La contraseña actual no es correcta.")
    conn.execute(
        "UPDATE usuario SET password_hash = ? WHERE id_usuario = ?",
        (generate_password_hash(new_pass), user_id),
    )
    conn.commit()
    conn.close()


def _update_user_email(user_id: int, email: str) -> None:
    correo = email.strip()[:150]
    if correo and not re.fullmatch(r"[^@]+@[^@]+\.[^@]+", correo):
        raise ValueError("Ingresa un correo electrónico válido.")
    conn = get_conn()
    conn.execute(
        "UPDATE usuario SET correo_usuario = ? WHERE id_usuario = ?",
        (correo or None, user_id),
    )
    conn.commit()
    conn.close()


def _parse_simulation_form(form) -> dict:
    capitalizacion_raw = form.get("capitalizacion", "").strip()
    capitalizacion = int(capitalizacion_raw) if capitalizacion_raw else None
    periodo_tasa = int(form.get("periodo_tasa", "7"))
    modalidad = form.get("modalidad", MODALIDAD_CONVENCIONAL).strip()
    cuota_balon_raw = form.get("cuota_balon_pct", "0").strip()
    cuota_balon_pct = float(cuota_balon_raw) if cuota_balon_raw else 0.0
    tipo_cambio_raw = form.get("tipo_cambio", "").strip()
    tipo_cambio = float(tipo_cambio_raw) if tipo_cambio_raw else None
    data = {
        "nombres_cliente": form["nombres_cliente"].strip()[:100],
        "apellidos_cliente": form["apellidos_cliente"].strip()[:100],
        "dni_cliente": form["dni_cliente"].strip(),
        "correo_cliente": form.get("correo_cliente", "").strip()[:150],
        "telefono_cliente": form.get("telefono_cliente", "").strip()[:20],
        "direccion_cliente": form.get("direccion_cliente", "").strip()[:200],
        "ingresos_mensuales": float(form["ingresos_mensuales"]),
        "marca_vehiculo": form["marca_vehiculo"].strip()[:50],
        "modelo_vehiculo": form["modelo_vehiculo"].strip()[:50],
        "precio_vehiculo": float(form["precio_vehiculo"]),
        "cuota_inicial_pct": float(form["cuota_inicial_pct"]),
        "moneda": form["moneda"],
        "tipo_tasa": form["tipo_tasa"],
        "tasa_interes": float(form["tasa_interes"]) / 100.0,
        "capitalizacion": capitalizacion,
        "periodo_tasa": periodo_tasa,
        "plazo_meses": int(form["plazo_meses"]),
        "periodo_gracia": form["periodo_gracia"],
        "meses_gracia": int(form.get("meses_gracia", 0)),
        "seguro_desgravamen": float(form["seguro_desgravamen"]) / 100.0,
        "seguro_vehicular": float(form["seguro_vehicular"]) / 100.0,
        "portes": float(form["portes"]),
        "fecha_desembolso": form.get("fecha_desembolso", "").strip(),
        "modalidad": modalidad,
        "cuota_balon_pct": cuota_balon_pct,
        "catalogo_id": form.get("catalogo_id", "").strip(),
        "gastos_notariales": float(form.get("gastos_notariales", 0) or 0),
        "gastos_registrales": float(form.get("gastos_registrales", 0) or 0),
        "costos_iniciales": float(form.get("costos_iniciales", 0) or 0),
        "tipo_cambio": tipo_cambio,
    }
    if not data["nombres_cliente"] or not data["apellidos_cliente"]:
        raise ValueError("Nombres y apellidos son obligatorios.")
    if not re.fullmatch(r"\d{8}", data["dni_cliente"]):
        raise ValueError("El DNI del cliente debe tener exactamente 8 dígitos.")
    if data["moneda"] not in ("Soles", "Dólares"):
        raise ValueError("Moneda inválida.")
    if data["tipo_tasa"] not in ("Efectiva", "Nominal"):
        raise ValueError("Tipo de tasa inválido.")
    if data["periodo_tasa"] < 0 or data["periodo_tasa"] > 7:
        data["periodo_tasa"] = 7
    if data["tipo_tasa"] == "Nominal" and (data["capitalizacion"] is None or data["capitalizacion"] < 1):
        raise ValueError("Para tasa nominal indica la capitalización (≥ 1).")
    if data["precio_vehiculo"] <= 0:
        raise ValueError("El precio del vehículo debe ser mayor a 0.")
    if not (0 <= data["cuota_inicial_pct"] < 100):
        raise ValueError("La cuota inicial debe estar entre 0 % y 99 %.")
    if data["tasa_interes"] <= 0:
        raise ValueError("La tasa de interés debe ser mayor a 0.")
    if data["plazo_meses"] < 1 or data["plazo_meses"] > 480:
        raise ValueError("El plazo debe estar entre 1 y 480 meses.")
    if data["periodo_gracia"] not in ("Ninguno", "Parcial", "Total"):
        raise ValueError("Periodo de gracia inválido.")
    if data["meses_gracia"] < 0 or data["meses_gracia"] >= data["plazo_meses"]:
        raise ValueError("Los meses de gracia deben ser menores al plazo total.")
    if data["ingresos_mensuales"] <= 0:
        raise ValueError("Los ingresos mensuales deben ser mayores a 0.")
    if data["modalidad"] not in (MODALIDAD_CONVENCIONAL, MODALIDAD_COMPRA_INTELIGENTE):
        raise ValueError("Modalidad inválida.")
    if data["modalidad"] == MODALIDAD_COMPRA_INTELIGENTE:
        if data["cuota_balon_pct"] <= 0 or data["cuota_balon_pct"] >= 100:
            raise ValueError("La cuota balón debe estar entre 0.01 % y 99 %.")
    elif data["cuota_balon_pct"] != 0:
        data["cuota_balon_pct"] = 0.0
    if data["gastos_notariales"] < 0 or data["gastos_registrales"] < 0 or data["costos_iniciales"] < 0:
        raise ValueError("Los gastos notariales, registrales e iniciales no pueden ser negativos.")
    if data["moneda"] == "Dólares":
        if not data["tipo_cambio"] or data["tipo_cambio"] <= 0:
            raise ValueError("Indica el tipo de cambio (soles por dólar) para créditos en dólares.")
    else:
        data["tipo_cambio"] = data["tipo_cambio"] if data["tipo_cambio"] and data["tipo_cambio"] > 0 else None
    if data["seguro_desgravamen"] < 0 or data["seguro_vehicular"] < 0 or data["portes"] < 0:
        raise ValueError("Seguros y portes no pueden ser negativos.")
    if not data["fecha_desembolso"] or not re.fullmatch(r"\d{4}-\d{2}-\d{2}", data["fecha_desembolso"]):
        raise ValueError("La fecha de desembolso es obligatoria (AAAA-MM-DD).")
    return data


@app.get("/")
def root():
    if "user_id" in session:
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))


@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        password2 = request.form.get("password_confirm", "").strip()
        dni = request.form.get("dni_usuario", "").strip()
        correo = request.form.get("correo_usuario", "").strip()[:150]
        if not username or not password:
            flash("Completa usuario y contraseña.")
            return render_template("register.html")
        if password != password2:
            flash("Las contraseñas no coinciden.")
            return render_template("register.html")
        if len(username) < 3 or len(username) > 50:
            flash("El usuario debe tener entre 3 y 50 caracteres.")
            return render_template("register.html")
        if len(password) < 6:
            flash("La contraseña debe tener al menos 6 caracteres.")
            return render_template("register.html")
        if not re.fullmatch(r"\d{8}", dni):
            flash("El DNI debe tener exactamente 8 dígitos.")
            return render_template("register.html")
        if correo and not re.fullmatch(r"[^@]+@[^@]+\.[^@]+", correo):
            flash("Ingresa un correo electrónico válido.")
            return render_template("register.html")

        if username.lower() == ADMIN_LOGIN:
            flash("Ese nombre de usuario está reservado.")
            return render_template("register.html")
        conn = get_conn()
        try:
            dup = conn.execute(
                "SELECT 1 FROM usuario WHERE usuario_login = ? OR dni_usuario = ?",
                (username, dni),
            ).fetchone()
            if dup:
                flash("Ese usuario o DNI ya está registrado.")
                return render_template("register.html")
            conn.execute(
                """
                INSERT INTO usuario (usuario_login, password_hash, dni_usuario, rol, correo_usuario)
                VALUES (?, ?, ?, ?, ?)
                """,
                (username, generate_password_hash(password), dni, ROL_USUARIO, correo or None),
            )
            conn.commit()
            flash("Cuenta creada. Ahora inicia sesión.")
            return redirect(url_for("login"))
        except Exception:
            flash("No se pudo registrar. Intenta con otro usuario o DNI.")
            return render_template("register.html")
        finally:
            conn.close()

    return render_template("register.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()

        conn = get_conn()
        user = conn.execute(
            "SELECT * FROM usuario WHERE usuario_login = ?", (username,)
        ).fetchone()
        conn.close()

        if user and check_password_hash(user["password_hash"], password):
            session["user_id"] = user["id_usuario"]
            session["username"] = user["usuario_login"]
            raw_dni = user["dni_usuario"] if "dni_usuario" in user.keys() else None
            session["dni"] = (raw_dni or "").strip() if raw_dni is not None else ""
            rol = user["rol"] if "rol" in user.keys() and user["rol"] else ROL_USUARIO
            session["rol"] = rol
            if rol == ROL_ADMIN:
                return redirect(url_for("admin_panel"))
            return redirect(url_for("dashboard"))
        flash("Credenciales inválidas.")
    return render_template("login.html")


@app.get("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


def _list_creditos(id_usuario=None, all_users: bool = False):
    conn = get_conn()
    if all_users:
        rows = conn.execute(
            """
            SELECT c.id_credito AS id, cl.nombre_cliente, cl.apellido_cliente,
                   v.marca_vehiculo, v.modelo_vehiculo, c.moneda, c.fecha_desembolso, c.created_at,
                   COALESCE(c.modalidad, 'Convencional') AS modalidad,
                   u.usuario_login
            FROM credito c
            JOIN cliente cl ON c.id_cliente = cl.id_cliente
            JOIN usuario u ON cl.id_usuario = u.id_usuario
            JOIN vehiculo v ON c.id_vehiculo = v.id_vehiculo
            ORDER BY c.id_credito DESC
            """
        ).fetchall()
    else:
        rows = conn.execute(
            """
            SELECT c.id_credito AS id, cl.nombre_cliente, cl.apellido_cliente,
                   v.marca_vehiculo, v.modelo_vehiculo, c.moneda, c.fecha_desembolso, c.created_at,
                   COALESCE(c.modalidad, 'Convencional') AS modalidad,
                   NULL AS usuario_login
            FROM credito c
            JOIN cliente cl ON c.id_cliente = cl.id_cliente
            JOIN vehiculo v ON c.id_vehiculo = v.id_vehiculo
            WHERE cl.id_usuario = ?
            ORDER BY c.id_credito DESC
            """,
            (id_usuario,),
        ).fetchall()
    conn.close()
    return rows


def _insert_cliente(conn, id_usuario: int, data: dict) -> int:
    cur = conn.execute(
        """
        INSERT INTO cliente (
            id_usuario, nombre_cliente, apellido_cliente, dni_cliente, correo_cliente,
            telefono_cliente, direccion_cliente, ingresos_mensuales
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            id_usuario,
            data["nombres_cliente"],
            data["apellidos_cliente"],
            data["dni_cliente"],
            data["correo_cliente"],
            data["telefono_cliente"],
            data["direccion_cliente"],
            data["ingresos_mensuales"],
        ),
    )
    return cur.lastrowid


def _parse_catalogo_id(raw) -> int | None:
    if raw in (None, ""):
        return None
    try:
        value = int(raw)
    except (TypeError, ValueError):
        return None
    return value if value > 0 else None


def _delete_credito(conn, credito_id: int) -> None:
    row = conn.execute(
        "SELECT id_cliente, id_vehiculo FROM credito WHERE id_credito = ?",
        (credito_id,),
    ).fetchone()
    if not row:
        return
    id_cliente = row["id_cliente"]
    id_vehiculo = row["id_vehiculo"]
    conn.execute("DELETE FROM cronograma_pago WHERE id_credito = ?", (credito_id,))
    conn.execute("DELETE FROM indicadores_financieros WHERE id_credito = ?", (credito_id,))
    conn.execute("DELETE FROM seguro WHERE id_credito = ?", (credito_id,))
    conn.execute("DELETE FROM credito WHERE id_credito = ?", (credito_id,))
    if not conn.execute(
        "SELECT 1 FROM credito WHERE id_cliente = ? LIMIT 1",
        (id_cliente,),
    ).fetchone():
        conn.execute("DELETE FROM cliente WHERE id_cliente = ?", (id_cliente,))
    if not conn.execute(
        "SELECT 1 FROM credito WHERE id_vehiculo = ? LIMIT 1",
        (id_vehiculo,),
    ).fetchone():
        conn.execute("DELETE FROM vehiculo WHERE id_vehiculo = ?", (id_vehiculo,))


def _credito_access_clause() -> tuple[str, list]:
    if _is_admin():
        return "", []
    return " AND cl.id_usuario = ?", [session["user_id"]]


def _indicadores_from_flujo(row) -> dict:
    """Recalcula VAN/TIR/TCEA desde flujo_json para evitar datos viejos en BD."""
    flujo_raw = _row_get(row, "flujo_json")
    tem = _num(_row_get(row, "tem"))
    if flujo_raw:
        try:
            flujo = json.loads(flujo_raw)
            if isinstance(flujo, list) and flujo:
                tir = calcular_tir(flujo)
                return {
                    "van": calcular_van(tem, flujo),
                    "tir": tir,
                    "tcea": calcular_tcea_desde_tir_mensual(tir),
                }
        except (json.JSONDecodeError, TypeError, ValueError):
            pass
    return {
        "van": _num(_row_get(row, "van")),
        "tir": _num(_row_get(row, "tir")),
        "tcea": _num(_row_get(row, "tcea")),
    }


def _van_costo_deudor(van: float) -> float:
    """Costo en valor presente (positivo) para mostrar al usuario."""
    return -van


def _delete_usuario_completo(target_id: int) -> None:
    if target_id == session.get("user_id"):
        raise ValueError("No puedes eliminar tu propia cuenta mientras estás conectado.")
    conn = get_conn()
    user = conn.execute(
        "SELECT id_usuario, usuario_login, rol FROM usuario WHERE id_usuario = ?",
        (target_id,),
    ).fetchone()
    if not user:
        conn.close()
        raise ValueError("Usuario no encontrado.")
    if user["rol"] == ROL_ADMIN:
        conn.close()
        raise ValueError("No se puede eliminar la cuenta de administrador.")
    creditos = conn.execute(
        """
        SELECT c.id_credito FROM credito c
        JOIN cliente cl ON c.id_cliente = cl.id_cliente
        WHERE cl.id_usuario = ?
        """,
        (target_id,),
    ).fetchall()
    for row in creditos:
        _delete_credito(conn, row["id_credito"])
    conn.execute("DELETE FROM cliente WHERE id_usuario = ?", (target_id,))
    conn.execute("DELETE FROM usuario WHERE id_usuario = ?", (target_id,))
    conn.commit()
    conn.close()


def _create_credito(uid: int, data: dict) -> int:
    result = build_schedule(
        precio_vehiculo=data["precio_vehiculo"],
        cuota_inicial_pct=data["cuota_inicial_pct"],
        tipo_tasa=data["tipo_tasa"],
        tasa_interes=data["tasa_interes"],
        plazo_meses=data["plazo_meses"],
        periodo_gracia=data["periodo_gracia"],
        meses_gracia=data["meses_gracia"],
        seguro_desgravamen=data["seguro_desgravamen"],
        seguro_vehicular=data["seguro_vehicular"],
        portes=data["portes"],
        capitalizacion=data["capitalizacion"],
        periodo_tasa=data["periodo_tasa"],
        modalidad=data["modalidad"],
        cuota_balon_pct=data["cuota_balon_pct"],
        gastos_notariales=data["gastos_notariales"],
        gastos_registrales=data["gastos_registrales"],
        costos_iniciales=data["costos_iniciales"],
    )

    conn = get_conn()
    cur = conn.cursor()
    id_cliente = _insert_cliente(conn, uid, data)
    catalogo_id = _parse_catalogo_id(data.get("catalogo_id"))

    cur.execute(
        "INSERT INTO vehiculo (id_catalogo, marca_vehiculo, modelo_vehiculo, precio_vehiculo) VALUES (?, ?, ?, ?)",
        (catalogo_id, data["marca_vehiculo"], data["modelo_vehiculo"], data["precio_vehiculo"]),
    )
    id_vehiculo = cur.lastrowid

    cur.execute(
        """
        INSERT INTO credito (
            id_cliente, id_vehiculo, moneda, tipo_tasa, tasa_interes, capitalizacion, periodo_tasa, plazo_meses,
            periodo_gracia, meses_gracia, cuota_inicial, fecha_desembolso, tem, cuota_base, total_financiado, flujo_json,
            modalidad, cuota_balon_pct, cuota_balon_monto, gastos_notariales, gastos_registrales, costos_iniciales, tipo_cambio
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            id_cliente,
            id_vehiculo,
            data["moneda"],
            data["tipo_tasa"],
            data["tasa_interes"],
            data["capitalizacion"],
            data["periodo_tasa"],
            data["plazo_meses"],
            data["periodo_gracia"],
            data["meses_gracia"],
            data["cuota_inicial_pct"],
            data["fecha_desembolso"],
            result.tem,
            result.cuota_base,
            result.saldo_inicial,
            json.dumps(result.flujo),
            result.modalidad,
            data["cuota_balon_pct"],
            result.cuota_balon_monto,
            data["gastos_notariales"],
            data["gastos_registrales"],
            data["costos_iniciales"],
            data["tipo_cambio"],
        ),
    )
    id_credito = cur.lastrowid

    cur.execute(
        "INSERT INTO seguro (id_credito, seguro_desgravamen, seguro_vehicular, portes) VALUES (?, ?, ?, ?)",
        (id_credito, data["seguro_desgravamen"], data["seguro_vehicular"], data["portes"]),
    )
    cur.execute(
        "INSERT INTO indicadores_financieros (id_credito, van, tir, tcea, tem) VALUES (?, ?, ?, ?, ?)",
        (id_credito, result.van, result.tir, result.tcea, result.tem),
    )
    for row in result.schedule:
        cur.execute(
            """
            INSERT INTO cronograma_pago (
                id_credito, numero_cuota, cuota_base, interes_periodo, amortizacion_periodo,
                saldo_pendiente, cuota_total, seguro_cuota, portes_cuota, cuota_balon
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                id_credito,
                row.periodo,
                result.cuota_base,
                row.interes,
                row.amortizacion,
                row.saldo_final,
                row.cuota_total,
                row.seguro,
                row.portes,
                row.cuota_balon,
            ),
        )
    conn.commit()
    conn.close()
    return id_credito


def _persist_credito_result(
    conn,
    *,
    id_cliente: int,
    id_vehiculo: int,
    id_credito: int,
    data: dict,
    result,
) -> None:
    conn.execute(
        """
        UPDATE cliente SET nombre_cliente=?, apellido_cliente=?, dni_cliente=?, correo_cliente=?, telefono_cliente=?,
            direccion_cliente=?, ingresos_mensuales=? WHERE id_cliente=?
        """,
        (
            data["nombres_cliente"],
            data["apellidos_cliente"],
            data["dni_cliente"],
            data["correo_cliente"],
            data["telefono_cliente"],
            data["direccion_cliente"],
            data["ingresos_mensuales"],
            id_cliente,
        ),
    )
    conn.execute(
        "UPDATE vehiculo SET id_catalogo=?, marca_vehiculo=?, modelo_vehiculo=?, precio_vehiculo=? WHERE id_vehiculo=?",
        (
            _parse_catalogo_id(data.get("catalogo_id")),
            data["marca_vehiculo"],
            data["modelo_vehiculo"],
            data["precio_vehiculo"],
            id_vehiculo,
        ),
    )
    conn.execute(
        """
        UPDATE credito SET
            moneda=?, tipo_tasa=?, tasa_interes=?, capitalizacion=?, periodo_tasa=?, plazo_meses=?,
            periodo_gracia=?, meses_gracia=?, cuota_inicial=?, fecha_desembolso=?, tem=?, cuota_base=?,
            total_financiado=?, flujo_json=?, modalidad=?, cuota_balon_pct=?, cuota_balon_monto=?,
            gastos_notariales=?, gastos_registrales=?, costos_iniciales=?, tipo_cambio=?
        WHERE id_credito=?
        """,
        (
            data["moneda"],
            data["tipo_tasa"],
            data["tasa_interes"],
            data["capitalizacion"],
            data["periodo_tasa"],
            data["plazo_meses"],
            data["periodo_gracia"],
            data["meses_gracia"],
            data["cuota_inicial_pct"],
            data["fecha_desembolso"],
            result.tem,
            result.cuota_base,
            result.saldo_inicial,
            json.dumps(result.flujo),
            result.modalidad,
            data["cuota_balon_pct"],
            result.cuota_balon_monto,
            data["gastos_notariales"],
            data["gastos_registrales"],
            data["costos_iniciales"],
            data["tipo_cambio"],
            id_credito,
        ),
    )
    conn.execute("DELETE FROM cronograma_pago WHERE id_credito = ?", (id_credito,))
    conn.execute("DELETE FROM indicadores_financieros WHERE id_credito = ?", (id_credito,))
    conn.execute("DELETE FROM seguro WHERE id_credito = ?", (id_credito,))
    conn.execute(
        "INSERT INTO seguro (id_credito, seguro_desgravamen, seguro_vehicular, portes) VALUES (?, ?, ?, ?)",
        (id_credito, data["seguro_desgravamen"], data["seguro_vehicular"], data["portes"]),
    )
    conn.execute(
        "INSERT INTO indicadores_financieros (id_credito, van, tir, tcea, tem) VALUES (?, ?, ?, ?, ?)",
        (id_credito, result.van, result.tir, result.tcea, result.tem),
    )
    for row in result.schedule:
        conn.execute(
            """
            INSERT INTO cronograma_pago (
                id_credito, numero_cuota, cuota_base, interes_periodo, amortizacion_periodo,
                saldo_pendiente, cuota_total, seguro_cuota, portes_cuota, cuota_balon
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                id_credito,
                row.periodo,
                result.cuota_base,
                row.interes,
                row.amortizacion,
                row.saldo_final,
                row.cuota_total,
                row.seguro,
                row.portes,
                row.cuota_balon,
            ),
        )


def _update_credito(credito_id: int, data: dict) -> int:
    result = build_schedule(
        precio_vehiculo=data["precio_vehiculo"],
        cuota_inicial_pct=data["cuota_inicial_pct"],
        tipo_tasa=data["tipo_tasa"],
        tasa_interes=data["tasa_interes"],
        plazo_meses=data["plazo_meses"],
        periodo_gracia=data["periodo_gracia"],
        meses_gracia=data["meses_gracia"],
        seguro_desgravamen=data["seguro_desgravamen"],
        seguro_vehicular=data["seguro_vehicular"],
        portes=data["portes"],
        capitalizacion=data["capitalizacion"],
        periodo_tasa=data["periodo_tasa"],
        modalidad=data["modalidad"],
        cuota_balon_pct=data["cuota_balon_pct"],
        gastos_notariales=data["gastos_notariales"],
        gastos_registrales=data["gastos_registrales"],
        costos_iniciales=data["costos_iniciales"],
    )
    conn = get_conn()
    extra, params = _credito_access_clause()
    row = conn.execute(
        f"""
        SELECT c.id_credito, c.id_cliente, c.id_vehiculo
        FROM credito c
        JOIN cliente cl ON c.id_cliente = cl.id_cliente
        WHERE c.id_credito = ?{extra}
        """,
        [credito_id, *params],
    ).fetchone()
    if not row:
        conn.close()
        raise ValueError("Simulación no encontrada o sin permiso para editarla.")
    _persist_credito_result(
        conn,
        id_cliente=row["id_cliente"],
        id_vehiculo=row["id_vehiculo"],
        id_credito=credito_id,
        data=data,
        result=result,
    )
    conn.commit()
    conn.close()
    return credito_id


def _load_credito_defaults(credito_id: int) -> dict | None:
    extra, params = _credito_access_clause()
    conn = get_conn()
    row = conn.execute(
        f"""
        SELECT c.*, cl.nombre_cliente, cl.apellido_cliente, cl.dni_cliente, cl.correo_cliente, cl.telefono_cliente,
               cl.direccion_cliente, cl.ingresos_mensuales,
               v.marca_vehiculo, v.modelo_vehiculo, v.precio_vehiculo, v.id_catalogo,
               s.seguro_desgravamen, s.seguro_vehicular, s.portes
        FROM credito c
        JOIN cliente cl ON c.id_cliente = cl.id_cliente
        JOIN vehiculo v ON c.id_vehiculo = v.id_vehiculo
        LEFT JOIN seguro s ON c.id_credito = s.id_credito
        WHERE c.id_credito = ?{extra}
        """,
        [credito_id, *params],
    ).fetchone()
    conn.close()
    if not row:
        return None
    catalogo_id = row["id_catalogo"] if row["id_catalogo"] else _match_catalogo_id(
        row["marca_vehiculo"], row["modelo_vehiculo"], row["precio_vehiculo"]
    )
    return {
        "nombres_cliente": row["nombre_cliente"],
        "apellidos_cliente": row["apellido_cliente"],
        "dni_cliente": row["dni_cliente"] or "",
        "correo_cliente": row["correo_cliente"] or "",
        "telefono_cliente": row["telefono_cliente"] or "",
        "direccion_cliente": row["direccion_cliente"] or "",
        "ingresos_mensuales": str(row["ingresos_mensuales"]),
        "marca_vehiculo": row["marca_vehiculo"],
        "modelo_vehiculo": row["modelo_vehiculo"],
        "precio_vehiculo": str(row["precio_vehiculo"]),
        "catalogo_id": str(catalogo_id) if catalogo_id else "",
        "cuota_inicial_pct": str(row["cuota_inicial"]),
        "moneda": row["moneda"],
        "modalidad": row["modalidad"] or MODALIDAD_CONVENCIONAL,
        "tipo_cambio": str(row["tipo_cambio"]) if row["tipo_cambio"] else "",
        "cuota_balon_pct": str(row["cuota_balon_pct"] or 0),
        "tipo_tasa": row["tipo_tasa"],
        "periodo_tasa": str(row["periodo_tasa"] if row["periodo_tasa"] is not None else 7),
        "tasa_interes": str(round(float(row["tasa_interes"]) * 100, 6)),
        "capitalizacion": str(row["capitalizacion"]) if row["capitalizacion"] is not None else "",
        "plazo_meses": str(row["plazo_meses"]),
        "fecha_desembolso": row["fecha_desembolso"] or "",
        "periodo_gracia": row["periodo_gracia"] or "Ninguno",
        "meses_gracia": str(row["meses_gracia"] or 0),
        "seguro_desgravamen": str(round(float(row["seguro_desgravamen"] or 0) * 100, 6)),
        "seguro_vehicular": str(round(float(row["seguro_vehicular"] or 0) * 100, 6)),
        "portes": str(row["portes"] or 0),
        "gastos_notariales": str(row["gastos_notariales"] or 0),
        "gastos_registrales": str(row["gastos_registrales"] or 0),
        "costos_iniciales": str(row["costos_iniciales"] or 0),
    }


def _render_wizard_page(defaults, *, edit_credito_id=None, selected_catalogo_id=None, flash_caso=None):
    if flash_caso and flash_caso in CASOS_PRUEBA:
        defaults.update(CASOS_PRUEBA[flash_caso])
        etiqueta = next((label for key, label in CASOS_PRUEBA_ORDEN if key == flash_caso), flash_caso.title())
        flash(f"Caso de prueba cargado: {etiqueta}")
    if selected_catalogo_id is None and defaults.get("catalogo_id"):
        try:
            selected_catalogo_id = int(defaults["catalogo_id"])
        except (TypeError, ValueError):
            selected_catalogo_id = None
    if selected_catalogo_id is None and defaults.get("marca_vehiculo") and defaults.get("modelo_vehiculo"):
        selected_catalogo_id = _match_catalogo_id(
            defaults.get("marca_vehiculo", ""),
            defaults.get("modelo_vehiculo", ""),
            defaults.get("precio_vehiculo"),
        )
        if selected_catalogo_id:
            defaults["catalogo_id"] = str(selected_catalogo_id)
    return render_template(
        "wizard.html",
        defaults=defaults,
        catalogo=_list_catalogo(),
        selected_catalogo_id=selected_catalogo_id,
        periodo_opciones=PERIODO_OPCIONES,
        active_nav="wizard",
        modalidades=(MODALIDAD_CONVENCIONAL, MODALIDAD_COMPRA_INTELIGENTE),
        casos_prueba=CASOS_PRUEBA,
        casos_prueba_orden=CASOS_PRUEBA_ORDEN,
        edit_credito_id=edit_credito_id,
    )


CATEGORIAS_CATALOGO = ("SUV", "Sedán", "Deportivo")
COMBUSTIBLES_CATALOGO = ("Gasolina", "Híbrido", "Eléctrico", "Diésel")
CONDICIONES_CATALOGO = ("Nuevo", "Seminuevo")


def _list_catalogo():
    conn = get_conn()
    rows = conn.execute(
        """
        SELECT id_catalogo, marca, modelo, anio, variante, categoria, combustible,
               condicion, descripcion, precio, imagen_url
        FROM catalogo_vehiculo
        ORDER BY categoria ASC, marca ASC, modelo ASC
        """
    ).fetchall()
    if not rows:
        from db import _ensure_catalogo

        _ensure_catalogo(conn)
        rows = conn.execute(
            """
            SELECT id_catalogo, marca, modelo, anio, variante, categoria, combustible,
                   condicion, descripcion, precio, imagen_url
            FROM catalogo_vehiculo
            ORDER BY categoria ASC, marca ASC, modelo ASC
            """
        ).fetchall()
    conn.close()
    return rows


def _get_catalogo_item(catalogo_id: int):
    conn = get_conn()
    row = conn.execute(
        "SELECT * FROM catalogo_vehiculo WHERE id_catalogo = ?", (catalogo_id,)
    ).fetchone()
    conn.close()
    return row


def _match_catalogo_id(marca: str, modelo: str, precio=None) -> int | None:
    conn = get_conn()
    modelo_base = modelo.split()[0] if modelo else ""
    row = conn.execute(
        """
        SELECT id_catalogo FROM catalogo_vehiculo
        WHERE marca = ? AND (modelo = ? OR modelo = ? OR ? LIKE modelo || '%')
        ORDER BY id_catalogo ASC LIMIT 1
        """,
        (marca.strip(), modelo.strip(), modelo_base, modelo.strip()),
    ).fetchone()
    if not row and precio:
        try:
            row = conn.execute(
                """
                SELECT id_catalogo FROM catalogo_vehiculo
                WHERE marca = ? AND ABS(precio - ?) < 0.01
                ORDER BY id_catalogo ASC LIMIT 1
                """,
                (marca.strip(), float(precio)),
            ).fetchone()
        except (TypeError, ValueError):
            pass
    conn.close()
    return row["id_catalogo"] if row else None


def _parse_catalogo_form(form) -> dict:
    marca = form.get("marca", "").strip()
    modelo = form.get("modelo", "").strip()
    variante = form.get("variante", "").strip()
    categoria = form.get("categoria", "").strip()
    combustible = form.get("combustible", "Gasolina").strip()
    condicion = form.get("condicion", "Nuevo").strip()
    descripcion = form.get("descripcion", "").strip()
    imagen_url = form.get("imagen_url", "").strip()
    try:
        anio = int(form.get("anio", "0"))
        precio = float(form.get("precio", "0"))
    except ValueError as exc:
        raise ValueError("Año y precio deben ser numéricos.") from exc
    if not marca or not modelo or not variante:
        raise ValueError("Marca, modelo y variante son obligatorios.")
    if categoria not in CATEGORIAS_CATALOGO:
        raise ValueError("Categoría inválida.")
    if anio < 1990 or anio > 2035:
        raise ValueError("Año del vehículo fuera de rango.")
    if precio <= 0:
        raise ValueError("El precio debe ser mayor a cero.")
    return {
        "marca": marca,
        "modelo": modelo,
        "anio": anio,
        "variante": variante,
        "categoria": categoria,
        "combustible": combustible,
        "condicion": condicion,
        "descripcion": descripcion,
        "precio": precio,
        "imagen_url": imagen_url or None,
    }


@app.get("/catalogo")
@login_required
def catalogo():
    return render_template(
        "catalogo.html",
        vehiculos=_list_catalogo(),
        categorias=CATEGORIAS_CATALOGO,
        combustibles=COMBUSTIBLES_CATALOGO,
        condiciones=CONDICIONES_CATALOGO,
        active_nav="catalogo",
    )


@app.post("/catalogo")
@admin_required
def catalogo_create():
    try:
        data = _parse_catalogo_form(request.form)
        conn = get_conn()
        conn.execute(
            """
            INSERT INTO catalogo_vehiculo (
                marca, modelo, anio, variante, categoria, combustible, condicion,
                descripcion, precio, imagen_url
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                data["marca"],
                data["modelo"],
                data["anio"],
                data["variante"],
                data["categoria"],
                data["combustible"],
                data["condicion"],
                data["descripcion"],
                data["precio"],
                data["imagen_url"],
            ),
        )
        conn.commit()
        conn.close()
        flash(f"Vehículo {data['marca']} {data['modelo']} agregado al catálogo.")
    except Exception as exc:
        flash(str(exc))
    return redirect(url_for("catalogo"))


@app.post("/catalogo/<int:catalogo_id>/editar")
@admin_required
def catalogo_update(catalogo_id: int):
    try:
        data = _parse_catalogo_form(request.form)
        conn = get_conn()
        row = conn.execute(
            "SELECT id_catalogo FROM catalogo_vehiculo WHERE id_catalogo = ?",
            (catalogo_id,),
        ).fetchone()
        if not row:
            conn.close()
            flash("Vehículo no encontrado.")
            return redirect(url_for("catalogo"))
        conn.execute(
            """
            UPDATE catalogo_vehiculo SET
                marca=?, modelo=?, anio=?, variante=?, categoria=?, combustible=?,
                condicion=?, descripcion=?, precio=?, imagen_url=?
            WHERE id_catalogo=?
            """,
            (
                data["marca"],
                data["modelo"],
                data["anio"],
                data["variante"],
                data["categoria"],
                data["combustible"],
                data["condicion"],
                data["descripcion"],
                data["precio"],
                data["imagen_url"],
                catalogo_id,
            ),
        )
        conn.commit()
        conn.close()
        flash("Vehículo actualizado.")
    except Exception as exc:
        flash(str(exc))
    return redirect(url_for("catalogo"))


@app.post("/catalogo/<int:catalogo_id>/eliminar")
@admin_required
def catalogo_delete(catalogo_id: int):
    conn = get_conn()
    row = conn.execute(
        "SELECT marca, modelo FROM catalogo_vehiculo WHERE id_catalogo = ?",
        (catalogo_id,),
    ).fetchone()
    if not row:
        conn.close()
        flash("Vehículo no encontrado.")
        return redirect(url_for("catalogo"))
    conn.execute("DELETE FROM catalogo_vehiculo WHERE id_catalogo = ?", (catalogo_id,))
    conn.commit()
    conn.close()
    flash(f"{row['marca']} {row['modelo']} eliminado del catálogo.")
    return redirect(url_for("catalogo"))


@app.get("/dashboard")
@login_required
def dashboard():
    return render_template(
        "dashboard.html",
        plans=_list_creditos(all_users=_is_admin()) if _is_admin() else _list_creditos(session["user_id"]),
        is_admin_view=_is_admin(),
        active_nav="dashboard",
    )


@app.route("/wizard", methods=["GET", "POST"])
@login_required
def wizard():
    if request.method == "POST":
        try:
            data = _parse_simulation_form(request.form)
            id_credito = _create_credito(session["user_id"], data)
            return redirect(url_for("plan_detail", credito_id=id_credito))
        except Exception as exc:
            flash(f"Error en cálculo o registro: {exc}")
    defaults = _get_wizard_defaults()
    caso = request.args.get("caso", "").strip().lower()
    catalogo_id = request.args.get("catalogo", type=int)
    selected_catalogo_id = catalogo_id
    if catalogo_id:
        item = _get_catalogo_item(catalogo_id)
        if item:
            defaults.update(
                {
                    "catalogo_id": str(catalogo_id),
                    "marca_vehiculo": item["marca"],
                    "modelo_vehiculo": f"{item['modelo']} {item['anio']}",
                    "precio_vehiculo": str(item["precio"]),
                }
            )
            flash(f"Vehículo cargado: {item['marca']} {item['modelo']}.")
        else:
            flash("Vehículo del catálogo no encontrado.")
    return _render_wizard_page(defaults, flash_caso=caso, selected_catalogo_id=selected_catalogo_id)


@app.route("/plans/<int:credito_id>/editar", methods=["GET", "POST"])
@login_required
def plan_edit(credito_id: int):
    if request.method == "POST":
        try:
            data = _parse_simulation_form(request.form)
            _update_credito(credito_id, data)
            flash("Simulación actualizada correctamente.")
            return redirect(url_for("plan_detail", credito_id=credito_id))
        except Exception as exc:
            flash(f"Error al actualizar: {exc}")
            defaults = _load_credito_defaults(credito_id) or _get_wizard_defaults()
            defaults.update({k: request.form.get(k, defaults.get(k, "")) for k in DEFAULTS_KEYS if k in request.form})
            return _render_wizard_page(defaults, edit_credito_id=credito_id)
    defaults = _load_credito_defaults(credito_id)
    if not defaults:
        flash("Simulación no encontrada o sin permiso para editarla.")
        return redirect(url_for("dashboard"))
    flash(f"Editando simulación #{credito_id}. Modifica los datos y guarda los cambios.")
    return _render_wizard_page(defaults, edit_credito_id=credito_id)


@app.route("/settings", methods=["GET", "POST"])
@login_required
def settings():
    if request.method == "POST":
        action = request.form.get("action", "password")
        try:
            if action == "email":
                _update_user_email(session["user_id"], request.form.get("correo_usuario", ""))
                flash("Correo actualizado.")
            else:
                _change_password(
                    session["user_id"],
                    request.form.get("current_password", ""),
                    request.form.get("new_password", ""),
                    request.form.get("new_password_confirm", ""),
                )
                flash("Contraseña actualizada.")
        except Exception as exc:
            flash(str(exc))
        return redirect(url_for("settings"))
    conn = get_conn()
    user_row = conn.execute(
        "SELECT usuario_login, dni_usuario, rol, correo_usuario FROM usuario WHERE id_usuario = ?",
        (session["user_id"],),
    ).fetchone()
    conn.close()
    return render_template(
        "settings.html",
        account=user_row,
        active_nav="settings",
    )


@app.get("/admin")
@admin_required
def admin_panel():
    conn = get_conn()
    stats = {
        "usuarios": conn.execute(
            "SELECT COUNT(*) AS n FROM usuario WHERE rol IS NULL OR rol != ?",
            (ROL_ADMIN,),
        ).fetchone()["n"],
        "simulaciones": conn.execute("SELECT COUNT(*) AS n FROM credito").fetchone()["n"],
    }
    por_modalidad = conn.execute(
        """
        SELECT COALESCE(modalidad, 'Convencional') AS modalidad, COUNT(*) AS total
        FROM credito GROUP BY COALESCE(modalidad, 'Convencional')
        ORDER BY total DESC
        """
    ).fetchall()
    por_moneda = conn.execute(
        "SELECT moneda, COUNT(*) AS total FROM credito GROUP BY moneda ORDER BY total DESC"
    ).fetchall()
    creditos = conn.execute(
        """
        SELECT c.id_credito AS id, u.usuario_login, cl.nombre_cliente, cl.apellido_cliente,
               v.marca_vehiculo, v.modelo_vehiculo, c.moneda,
               COALESCE(c.modalidad, 'Convencional') AS modalidad, c.created_at
        FROM credito c
        JOIN cliente cl ON c.id_cliente = cl.id_cliente
        JOIN usuario u ON cl.id_usuario = u.id_usuario
        JOIN vehiculo v ON c.id_vehiculo = v.id_vehiculo
        ORDER BY c.id_credito DESC
        """
    ).fetchall()
    usuarios = conn.execute(
        """
        SELECT u.id_usuario, u.usuario_login, u.dni_usuario, u.created_at,
               COALESCE(u.rol, 'usuario') AS rol,
               (SELECT COUNT(*) FROM credito c
                JOIN cliente cl ON c.id_cliente = cl.id_cliente
                WHERE cl.id_usuario = u.id_usuario) AS num_creditos
        FROM usuario u
        ORDER BY u.id_usuario ASC
        """
    ).fetchall()
    conn.close()
    return render_template(
        "admin.html",
        stats=stats,
        por_modalidad=por_modalidad,
        por_moneda=por_moneda,
        creditos=creditos,
        usuarios=usuarios,
        active_nav="admin",
    )


@app.post("/admin/catalogo/reseed")
@admin_required
def admin_reseed_catalogo():
    from db import _ensure_catalogo

    conn = get_conn()
    _ensure_catalogo(conn)
    conn.close()
    flash(f"Catálogo restaurado: {catalogo_count()} vehículos.")
    return redirect(url_for("admin_panel"))


@app.post("/admin/users/<int:user_id>/delete")
@admin_required
def admin_delete_user(user_id: int):
    try:
        _delete_usuario_completo(user_id)
        flash("Usuario eliminado.")
    except Exception as exc:
        flash(str(exc))
    return redirect(url_for("admin_panel"))


@app.post("/admin/plans/<int:credito_id>/delete")
@admin_required
def admin_delete_plan(credito_id: int):
    try:
        conn = get_conn()
        row = conn.execute("SELECT id_credito FROM credito WHERE id_credito = ?", (credito_id,)).fetchone()
        if not row:
            conn.close()
            flash("Simulación no encontrada.")
            return redirect(url_for("admin_panel"))
        _delete_credito(conn, credito_id)
        conn.commit()
        conn.close()
        flash("Simulación eliminada.")
    except Exception as exc:
        flash(f"No se pudo eliminar la simulación: {exc}")
    return redirect(url_for("admin_panel"))


@app.get("/plans/<int:credito_id>")
@login_required
def plan_detail(credito_id: int):
    extra, params = _credito_access_clause()
    conn = get_conn()
    row = conn.execute(
        f"""
        SELECT c.*, cl.nombre_cliente, cl.apellido_cliente, cl.dni_cliente,
               v.marca_vehiculo, v.modelo_vehiculo, i.van, i.tir, i.tcea,
               s.seguro_desgravamen, s.seguro_vehicular, s.portes AS portes_cuota
        FROM credito c
        JOIN cliente cl ON c.id_cliente = cl.id_cliente
        JOIN vehiculo v ON c.id_vehiculo = v.id_vehiculo
        LEFT JOIN indicadores_financieros i ON c.id_credito = i.id_credito
        LEFT JOIN seguro s ON c.id_credito = s.id_credito
        WHERE c.id_credito = ?{extra}
        """,
        [credito_id, *params],
    ).fetchone()
    schedule = conn.execute(
        """
        SELECT numero_cuota AS periodo, cuota_base, interes_periodo AS interes, amortizacion_periodo AS amortizacion,
               saldo_pendiente AS saldo_final, cuota_total, seguro_cuota AS seguro, portes_cuota AS portes,
               COALESCE(cuota_balon, 0) AS cuota_balon
        FROM cronograma_pago WHERE id_credito = ? ORDER BY numero_cuota ASC
        """,
        (credito_id,),
    ).fetchall()
    conn.close()
    if not row:
        flash("Crédito no encontrado.")
        return redirect(url_for("dashboard"))

    periodo_tasa = _row_get(row, "periodo_tasa", 7)
    cap = _row_get(row, "capitalizacion")
    total_pagar = sum(_num(r["cuota_total"]) for r in schedule)
    sym = "S/" if _row_get(row, "moneda") == "Soles" else "$"

    modalidad = _row_get(row, "modalidad") or MODALIDAD_CONVENCIONAL
    cuota_balon_monto = _num(_row_get(row, "cuota_balon_monto"))
    gastos_not = _num(_row_get(row, "gastos_notariales"))
    gastos_reg = _num(_row_get(row, "gastos_registrales"))
    costos_ini = _num(_row_get(row, "costos_iniciales"))
    tipo_cambio_raw = _row_get(row, "tipo_cambio")
    tipo_cambio = _num(tipo_cambio_raw, default=0.0) if tipo_cambio_raw not in (None, "") else None

    indicadores = _indicadores_from_flujo(row)
    plan_view = {
        "id_credito": _row_get(row, "id_credito"),
        "nombre_cliente": _row_get(row, "nombre_cliente", ""),
        "apellido_cliente": _row_get(row, "apellido_cliente", ""),
        "dni_cliente": _row_get(row, "dni_cliente", ""),
        "marca_vehiculo": _row_get(row, "marca_vehiculo", ""),
        "modelo_vehiculo": _row_get(row, "modelo_vehiculo", ""),
        "moneda": _row_get(row, "moneda", "Soles"),
        "tipo_tasa": _row_get(row, "tipo_tasa", "Efectiva"),
        "fecha_desembolso": _row_get(row, "fecha_desembolso", ""),
        "periodo_gracia": _row_get(row, "periodo_gracia", "Ninguno"),
        "meses_gracia": _row_get(row, "meses_gracia", 0),
        "cuota_base": _num(_row_get(row, "cuota_base")),
        "tasa_interes": _num(_row_get(row, "tasa_interes")),
        "total_financiado": _num(_row_get(row, "total_financiado")),
        "tem": _num(_row_get(row, "tem")),
        "van": indicadores["van"],
        "van_costo": _van_costo_deudor(indicadores["van"]),
        "tir": indicadores["tir"],
        "tcea": indicadores["tcea"],
    }
    schedule_view = [
        {
            "periodo": int(_num(r["periodo"], 0)),
            "cuota_total": _num(r["cuota_total"]),
            "interes": _num(r["interes"]),
            "amortizacion": _num(r["amortizacion"]),
            "saldo_final": _num(r["saldo_final"]),
            "seguro": _num(r["seguro"]),
            "portes": _num(r["portes"]),
            "cuota_balon": _num(r["cuota_balon"]),
        }
        for r in schedule
    ]

    total_intereses = sum(r["interes"] for r in schedule_view)
    total_seguros = sum(r["seguro"] for r in schedule_view)
    total_portes = sum(r["portes"] for r in schedule_view)
    total_amortizacion = sum(r["amortizacion"] for r in schedule_view)
    seguro_desgravamen = _num(_row_get(row, "seguro_desgravamen"))
    seguro_vehicular = _num(_row_get(row, "seguro_vehicular"))
    portes_cuota = _num(_row_get(row, "portes_cuota"))

    transparencia_sbs = {
        "tcea_pct": plan_view["tcea"] * 100,
        "tea_pct": plan_view["tasa_interes"] * 100,
        "tem_pct": plan_view["tem"] * 100,
        "capital_financiado": plan_view["total_financiado"],
        "gastos_iniciales": gastos_not + gastos_reg + costos_ini,
        "total_intereses": total_intereses,
        "total_seguros": total_seguros,
        "total_portes": total_portes,
        "total_amortizacion": total_amortizacion,
        "total_a_pagar": total_pagar,
        "seguro_desgravamen_pct": seguro_desgravamen * 100,
        "seguro_vehicular_pct": seguro_vehicular * 100,
        "portes_mensual": portes_cuota,
        "num_cuotas": len(schedule_view),
    }

    return render_template(
        "plan_detail.html",
        plan=plan_view,
        schedule=schedule_view,
        periodo_tasa_etiqueta=etiqueta_periodo_tasa(periodo_tasa),
        capitalizacion_etiqueta=etiqueta_capitalizacion(cap),
        total_pagar=total_pagar,
        sym=sym,
        modalidad=modalidad,
        cuota_balon_monto=cuota_balon_monto,
        gastos_totales=gastos_not + gastos_reg + costos_ini,
        gastos_notariales=gastos_not,
        gastos_registrales=gastos_reg,
        costos_iniciales=costos_ini,
        tipo_cambio=tipo_cambio,
        transparencia_sbs=transparencia_sbs,
        chart_labels=[r["periodo"] for r in schedule_view],
        chart_intereses=[r["interes"] for r in schedule_view],
        chart_amort=[r["amortizacion"] for r in schedule_view],
        active_nav="admin" if _is_admin() else "dashboard",
    )


@app.post("/plans/<int:credito_id>/delete")
@login_required
def plan_delete(credito_id: int):
    extra, params = _credito_access_clause()
    conn = get_conn()
    row = conn.execute(
        f"""
        SELECT c.id_credito FROM credito c
        JOIN cliente cl ON c.id_cliente = cl.id_cliente
        WHERE c.id_credito = ?{extra}
        """,
        [credito_id, *params],
    ).fetchone()
    if not row:
        conn.close()
        flash("Simulación no encontrada.")
        return redirect(url_for("dashboard"))
    _delete_credito(conn, credito_id)
    conn.commit()
    conn.close()
    flash("Simulación eliminada.")
    redirect_to = url_for("admin_panel") if _is_admin() and request.referrer and "/admin" in request.referrer else url_for("dashboard")
    return redirect(redirect_to)


def _build_cronograma_xlsx(credito_id: int, meta: dict, schedule_rows) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Cronograma"

    title_font = Font(bold=True, size=14, color="9F1239")
    label_font = Font(bold=True)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="9F1239")
    total_fill = PatternFill(fill_type="solid", fgColor="F3F4F6")
    center = Alignment(horizontal="center", vertical="center")

    ws["A1"] = "FinanCuota — Cronograma de pagos"
    ws["A1"].font = title_font

    sym = "S/" if meta.get("moneda") == "Soles" else "USD"
    info_rows = [
        ("Credito N°", str(credito_id)),
        ("Cliente", meta.get("cliente", "")),
        ("Vehiculo", meta.get("vehiculo", "")),
        ("Modalidad", meta.get("modalidad", "")),
        ("Moneda", meta.get("moneda", "")),
        ("Capital financiado", f"{sym} {meta.get('capital', 0):,.2f}"),
        ("Plazo", f"{meta.get('plazo', 0)} meses"),
    ]
    row_idx = 3
    for label, value in info_rows:
        ws.cell(row=row_idx, column=1, value=label).font = label_font
        ws.cell(row=row_idx, column=2, value=value)
        row_idx += 1

    row_idx += 1
    header_row = row_idx
    headers = [
        "N°",
        "Cuota Total",
        "Interes",
        "Amortizacion",
        "Cuota Balon",
        "Seguro",
        "Portes",
        "Saldo Final",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    money_cols = {2, 3, 4, 5, 6, 7, 8}
    data_start = header_row + 1
    for sched in schedule_rows:
        row_idx += 1
        values = [
            int(_num(_row_get(sched, "numero_cuota"), 0)),
            _num(_row_get(sched, "cuota_total")),
            _num(_row_get(sched, "interes_periodo")),
            _num(_row_get(sched, "amortizacion_periodo")),
            _num(_row_get(sched, "cuota_balon")),
            _num(_row_get(sched, "seguro_cuota")),
            _num(_row_get(sched, "portes_cuota")),
            _num(_row_get(sched, "saldo_pendiente")),
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            if col in money_cols:
                cell.number_format = "#,##0.00"
            if col == 1:
                cell.alignment = center

    total_row = row_idx + 1
    ws.cell(row=total_row, column=1, value="TOTAL").font = label_font
    ws.cell(row=total_row, column=1).fill = total_fill
    total_cell = ws.cell(row=total_row, column=2, value=f"=SUM(B{data_start}:B{row_idx})")
    total_cell.font = label_font
    total_cell.fill = total_fill
    total_cell.number_format = "#,##0.00"

    widths = (6, 14, 12, 14, 14, 12, 10, 14)
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = ws.cell(row=data_start, column=1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@app.get("/plans/<int:credito_id>/csv")
@login_required
def plan_csv(credito_id: int):
    extra, params = _credito_access_clause()
    conn = get_conn()
    row = conn.execute(
        f"""
        SELECT c.id_credito, c.modalidad, c.moneda, c.plazo_meses, c.total_financiado,
               cl.nombre_cliente, cl.apellido_cliente,
               v.marca_vehiculo, v.modelo_vehiculo
        FROM credito c
        JOIN cliente cl ON c.id_cliente = cl.id_cliente
        JOIN vehiculo v ON c.id_vehiculo = v.id_vehiculo
        WHERE c.id_credito = ?{extra}
        """,
        [credito_id, *params],
    ).fetchone()
    schedule = conn.execute(
        """
        SELECT numero_cuota, cuota_total, interes_periodo, amortizacion_periodo,
               seguro_cuota, portes_cuota, saldo_pendiente, COALESCE(cuota_balon, 0) AS cuota_balon
        FROM cronograma_pago WHERE id_credito = ? ORDER BY numero_cuota
        """,
        (credito_id,),
    ).fetchall()
    conn.close()
    if not row:
        flash("Crédito no encontrado.")
        return redirect(url_for("dashboard"))

    meta = {
        "cliente": f"{_row_get(row, 'nombre_cliente', '')} {_row_get(row, 'apellido_cliente', '')}".strip(),
        "vehiculo": f"{_row_get(row, 'marca_vehiculo', '')} {_row_get(row, 'modelo_vehiculo', '')}".strip(),
        "modalidad": _row_get(row, "modalidad") or MODALIDAD_CONVENCIONAL,
        "moneda": _row_get(row, "moneda") or "Soles",
        "plazo": _row_get(row, "plazo_meses"),
        "capital": _num(_row_get(row, "total_financiado")),
    }
    xlsx_bytes = _build_cronograma_xlsx(credito_id, meta, schedule)
    filename = f"credito_{credito_id}_cronograma.xlsx"
    return Response(
        xlsx_bytes,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("FLASK_DEBUG", "1") == "1"
    app.run(debug=debug, host="0.0.0.0", port=port)
